import os
import bottlenose
import time
import traceback
import uuid

from bs4 import BeautifulSoup
from pymongo import MongoClient
from openpyxl import load_workbook
from json import dumps, loads

client = MongoClient(os.environ['DB_CONFIG']).stylens
amazon = bottlenose.Amazon(AWSAccessKeyId=os.environ['AWS_ACCESS_KEY_ID'],
                           AWSSecretAccessKey=os.environ['AWS_SECRET_ACCESS_KEY'],
                           AssociateTag=os.environ['AWS_ASSOCIATE_TAG'])

ROOT = os.path.join(os.path.dirname(os.path.realpath(__file__)), os.pardir)
location = os.path.join(ROOT, 'res/')

if __name__ == '__main__':
    if not os.path.exists(os.path.join(ROOT, 'temp')):
        os.mkdir(os.path.join(ROOT, 'temp'))

    if not os.path.exists(os.path.join(ROOT, 'fail')):
        os.mkdir(os.path.join(ROOT, 'fail'))

    for directory in os.listdir(location):
        print(directory)
        temp_file = os.path.join(ROOT, 'temp/%s.txt' % uuid.uuid4())
        fail_file = os.path.join(ROOT, 'fail/%s.txt' % uuid.uuid4())

        wb = load_workbook(filename=os.path.join(location, directory), read_only=True)
        with open(temp_file, 'w') as f:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                column = []
                for idx, row in enumerate(ws.rows):
                    product = {}
                    if idx == 0:
                        for cell in row:
                            column.append(cell.value)
                    else:
                        for cell_idx, cell in enumerate(row):
                            product[column[cell_idx]] = cell.value

                        if product['Category'] == 'apparel':
                            f.write(dumps(product) + '\n')

        with open(fail_file, 'w') as f2:
            with open(temp_file, 'r') as f1:
                while True:
                    line = f1.readline()

                    if not line:
                        break

                    parsed_product = loads(line)

                    try:
                        response = amazon.ItemLookup(ItemId=parsed_product['ASIN'],
                                                     ResponseGroup='Accessories,BrowseNodes,EditorialReview,Images,ItemAttributes,ItemIds,Large,Medium,OfferFull,Offers,PromotionSummary,OfferSummary,RelatedItems,Reviews,SalesRank,Similarities,Small,Tracks,VariationImages,Variations,VariationSummary')

                        soup = BeautifulSoup(response, "xml")
                        items = soup.find('Items')
                        item_list = items.find_all('Item')
                        for item in item_list:
                            parsed_product['DetailPageURL'] = item.DetailPageURL.text
                            if item.Binding:
                                parsed_product['Binding'] = item.Binding.text
                            if item.Department:
                                parsed_product['Department'] = item.Department.text
                            features = item.find_all('Feature')
                            parsed_product['Feature'] = ';'.join([feature.text for feature in features])

                            if item.Label:
                                parsed_product['Label'] = item.Label.text
                            if item.Manufacturer:
                                parsed_product['Manufacturer'] = item.Manufacturer.text

                            if item.Model:
                                parsed_product['Model'] = item.Model.text

                            parsed_product['ProductGroup'] = item.ProductGroup.text
                            parsed_product['ProductTypeName'] = item.ProductTypeName.text

                            if item.Publisher:
                                parsed_product['Publisher'] = item.Publisher.text

                            if item.Studio:
                                parsed_product['Studio'] = item.Studio.text

                            parsed_product['Title'] = item.Title.text

                            item_attr = item.find('ItemAttributes')

                            if item_attr:
                                if item_attr.Brand:
                                    parsed_product['Brand'] = item_attr.Brand.text

                                if item_attr.Color:
                                    parsed_product['Color'] = item_attr.Color.text

                                if item_attr.Label:
                                    parsed_product['Label'] = item_attr.Label.text

                                if item_attr.Manufacturer:
                                    parsed_product['Manufacturer'] = item_attr.Manufacturer.text

                                if item_attr.Model:
                                    parsed_product['Model'] = item_attr.Model.text

                                if item_attr.MPN:
                                    parsed_product['MPN'] = item_attr.MPN.text

                                if item_attr.Size:
                                    parsed_product['Size'] = item_attr.Size.text

                                if item_attr.Studio:
                                    parsed_product['Studio'] = item_attr.Studio.text

                                parsed_product['ItemDimensions'] = {}
                                if item_attr.ItemDimensions:
                                    for attr in item_attr.ItemDimensions:
                                        parsed_product['ItemDimensions'][attr.name] = attr.text

                        browse_nodes = soup.find('BrowseNodes')
                        browse_nodes_list = []
                        for browse_node in browse_nodes:
                            browse_node_list = []
                            if browse_node.Name:
                                browse_node_list.append(browse_node.Name.text)

                            for children in browse_node.find_all('Children'):
                                children_list = []
                                for name in children.find_all('Name'):
                                    children_list.append(name.text)
                                browse_node_list.append(children_list)

                            for ancestors in browse_node.find('Ancestors'):
                                for name in ancestors.find_all('Name'):
                                    browse_node_list.append(name.text)

                            browse_nodes_list.append(browse_node_list)
                        parsed_product['BrowseNodes'] = browse_nodes_list
                        parsed_product['OfferSummary'] = {}

                        offers_summary = soup.find('OfferSummary')

                        if offers_summary.LowestNewPrice:
                            if offers_summary.LowestNewPrice.FormattedPrice:
                                parsed_product['OfferSummary'][
                                    'LowestNewPrice'] = offers_summary.LowestNewPrice.FormattedPrice.text

                        if offers_summary.TotalNew:
                            parsed_product['OfferSummary']['TotalNew'] = offers_summary.TotalNew.text

                        if offers_summary.TotalUsed:
                            parsed_product['OfferSummary']['TotalUsed'] = offers_summary.TotalUsed.text

                        if offers_summary.TotalCollectible:
                            parsed_product['OfferSummary']['TotalCollectible'] = offers_summary.TotalCollectible.text

                        if offers_summary.TotalRefurbished:
                            parsed_product['OfferSummary']['TotalRefurbished'] = offers_summary.TotalRefurbished.text

                        offers = soup.find_all('Offer')

                        offers_list = []
                        for offer in offers:
                            offer_dic = {
                                'Condition': []
                            }

                            if offer.Merchant:
                                if offer.Merchant.Name:
                                    offer_dic['Merchant'] = offer.Merchant.Name.text

                            if offer.Price.FormattedPrice:
                                offer_dic['Price'] = offer.Price.FormattedPrice.text

                            for condition in offer.find_all('Condition'):
                                offer_dic['Condition'].append(condition.text)

                            offers_list.append(offer_dic)

                        parsed_product['Offer'] = offers_list
                        client.top_selling.insert_one(parsed_product)
                        time.sleep(2)

                    except Exception as ex:
                        f2.write(dumps(parsed_product) + '\n')
                        traceback.print_exc()
